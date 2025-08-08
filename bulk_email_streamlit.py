import streamlit as st
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import re
import openpyxl
import os
from datetime import datetime
import tempfile
import io
import time

# Initialize session state for login
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.query_params.clear()  # Replace st.experimental_set_query_params with st.query_params.clear()

# Valid users (use Streamlit secrets for deployment)
VALID_USERS = {
    "admin": "james",  # Replace with secure credentials
    "user1": "james"
}

# Email accounts (use Streamlit secrets for deployment)
EMAIL_ACCOUNTS = {
    "commercialcleaningbest@gmail.com": "uftl wdaz frbb cgda",
    "commercialcleaningm@gmail.com": "pzbp przh rmul sspz",
    "melbournecommercialcleaning2@gmail.com": "wusv clda ogbr knun",
    "ccleaning82@gmail.com": "rbte bvww wvde ywij",
    "yourgmail5@gmail.com": "your_app_password5"
}

# Basic email validation using regex
def validate_email(email):
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

# Send email using Gmail SMTP with HTML
def send_email(sender, password, recipient, name, subject, message_html, attachment_path=None):
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = f'"Sparkle Commercial Cleaning" <{sender}>'
    msg['To'] = recipient

    html_content = f"""
    <html>
    <body>
        <p>Dear {name},</p>
        {message_html}
    </body>
    </html>
    """
    msg.attach(MIMEText(html_content, 'html'))

    if attachment_path and os.path.exists(attachment_path):
        try:
            with open(attachment_path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename={os.path.basename(attachment_path)}'
            )
            msg.attach(part)
        except Exception as e:
            print(f"Failed to attach file for {recipient}: {e}")
            return False

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender, password)
        server.sendmail(sender, recipient, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"Failed to send to {recipient}: {e}")
        return False

# Load recipients from Excel
def load_recipients(excel_file):
    try:
        workbook = openpyxl.load_workbook(excel_file, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if headers != ['Name', 'Email']:
            workbook.close()
            return None, "Excel file must have 'Name' and 'Email' columns in the first row!"
        recipients = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, email = row[0], row[1]
            if name and email:
                recipients.append((str(name).strip(), str(email).strip()))
        workbook.close()  # Explicitly close the workbook
        return recipients, None
    except Exception as e:
        return None, f"Failed to read Excel file: {str(e)}"

# Save delivery report to Excel
def save_delivery_report(delivery_report):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Delivery Report"
    sheet.append(["Email", "Status", "Timestamp"])
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 2025-08-08 22:05:00
    for email in delivery_report["sent"]:
        sheet.append([email, "Sent", timestamp])
    for email in delivery_report["failed"]:
        sheet.append([email, "Failed", timestamp])
    
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"delivery_report_{timestamp.replace(' ', '_').replace(':', '')}.xlsx"
    return buffer, filename

# Safe file deletion with retries
def safe_unlink(file_path):
    max_attempts = 3
    for attempt in range(max_attempts):
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
            return True
        except PermissionError as e:
            if attempt < max_attempts - 1:
                time.sleep(1)
                continue
            print(f"Failed to delete {file_path}: {e}")
            return False

# Login page
def login_page():
    st.title("Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        if username in VALID_USERS and VALID_USERS[username] == password:
            st.session_state.logged_in = True
            st.success("Login successful!")
            st.rerun()
        else:
            st.error("Invalid username or password")
    st.write(f"Debug: logged_in = {st.session_state.logged_in}")  # Debug statement

# Main app content
def main_app():
    st.image(
        "https://sparklecommercialcleaning.com.au/wp-content/uploads/2023/05/cropped-sparkle_logo-1.png",
        width=200,
        caption=""
    )
    st.markdown("<h1 style='text-align: center; color: #333;'>Sparkle Commercial Cleaning</h1>", unsafe_allow_html=True)
    st.markdown("<h2 style='text-align: center; color: #555;'>Bulk Email Sender</h2>", unsafe_allow_html=True)

    sender = st.selectbox("Select Sender Email:", list(EMAIL_ACCOUNTS.keys()))
    subject = st.text_input("Subject:", key="subject")
    message_html = st.text_area("Message (Raw HTML, will add 'Dear [Name]' automatically):", height=150, key="message")
    excel_file = st.file_uploader("Excel File (.xlsx):", type=["xlsx"], key="excel")
    attachment = st.file_uploader("Attachment (optional):", type=["pdf", "jpg", "png", "doc", "docx", "txt"], key="attachment")

    if st.button("Send Emails"):
        if not excel_file:
            st.error("Please upload an Excel file.")
            return

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            tmp_excel.write(excel_file.read())
            excel_path = tmp_excel.name

        attachment_path = None
        if attachment:
            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(attachment.name)[1]) as tmp_attachment:
                tmp_attachment.write(attachment.read())
                attachment_path = tmp_attachment.name

        recipients, error = load_recipients(excel_path)
        if error:
            st.error(error)
            safe_unlink(excel_path)
            if attachment_path:
                safe_unlink(attachment_path)
            return

        if len(recipients) > 500:
            st.warning("Recipient list exceeds 500 emails. Gmail may block your account.")
            safe_unlink(excel_path)
            if attachment_path:
                safe_unlink(attachment_path)
            return

        valid_recipients = [(name, email) for name, email in recipients if validate_email(email)]
        if not valid_recipients:
            st.error("No valid emails found.")
            safe_unlink(excel_path)
            if attachment_path:
                safe_unlink(attachment_path)
            return

        progress_bar = st.progress(0)
        status_text = st.empty()
        total_emails = len(valid_recipients)
        delivery_report = {"sent": [], "failed": []}

        for i, (name, email) in enumerate(valid_recipients, 1):
            success = send_email(sender, EMAIL_ACCOUNTS[sender], email, name, subject, message_html, attachment_path)
            if success:
                delivery_report["sent"].append(email)
            else:
                delivery_report["failed"].append(email)
            
            progress = i / total_emails
            progress_bar.progress(progress)
            status_text.text(f"Progress: {i}/{total_emails} emails sent")

        safe_unlink(excel_path)
        if attachment_path:
            safe_unlink(attachment_path)

        st.info(f"Delivery Report: Sent: {len(delivery_report['sent'])}, Failed: {len(delivery_report['failed'])}")
        report_buffer, report_filename = save_delivery_report(delivery_report)
        st.download_button(
            label="Download Delivery Report",
            data=report_buffer,
            file_name=report_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Main app logic
def main():
    st.write(f"Debug: Initial logged_in = {st.session_state.logged_in}")  # Debug initial state
    if not st.session_state.logged_in:
        login_page()
    else:
        main_app()

if __name__ == "__main__":
    main()
